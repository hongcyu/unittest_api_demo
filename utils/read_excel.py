# -*-coding: utf8 -*-

'''
read_excel.py
读取excel表单内的数据，通过配置文件实现测试用例的配置,只能是xslx
新建excel文件时不可以在项目内部直接建,否则会报错
自动识别行列
'''
from read_config import Read_config
from openpyxl import load_workbook
import init_path

class Read_the_excel(object):
    #读取excel的内容
    @staticmethod
    def get_data(data_path,config_path):
        wb = load_workbook(data_path)   #创建workbook对象
        modes = eval(Read_config.get_config(config_path,'MODE','mode')) #清除注释
        
        test_data = []
        for mode in modes:
            sheet = wb[mode]
            if modes[mode] == 'all':
                for row in range(2,sheet.max_row +1):
                    tmp_data = dict()
                    tmp_data['sheet_name'] = mode
                    for column in range(1,sheet.max_column+1):
                        tmp_data[sheet.cell(1,column).value] = sheet.cell(row,column).value
                    test_data.append(tmp_data)
            else:
                for case_id in modes[mode]:
                    tmp_data = dict()
                    tmp_data['sheet_name'] = mode
                    for column in range(1,sheet.max_column+1):
                        tmp_data[sheet.cell(1,column).value] = sheet.cell(case_id + 1,column).value
                    test_data.append(tmp_data)

        return test_data
    
    #写入执行结果
    @staticmethod
    def write_excel(data_path,sheet_name,row,result,TestResult):
        pass

if __name__ == "__main__":
    pass
    # data_path = init_path.test_case_path
    # config_path = init_path.config_path
    # excel = Read_the_excel.get_data(data_path,config_path)
    # print(excel)
